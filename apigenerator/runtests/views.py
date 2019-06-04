# -*- coding: utf-8 -*-
"""
Views of runtests app
"""

import json
import urllib
import re
import numpy as np
import pandas as pd

from openpyxl import load_workbook

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import PermissionDenied
from django.http import JsonResponse
from django.urls import reverse_lazy, reverse
from django.views.generic import TemplateView
from django.views.generic.edit import CreateView, UpdateView, DeleteView

from obp.api import API, APIError
import logging
from .forms import *
from .models import *

LOGGER = logging.getLogger(__name__)
# TODO: These have to map to attributes of models.TestConfiguration
URLPATH_DEFAULT = [
    '3.1.0',
    'Test', 'Test', '1',
    '1', '1', '1', '1',
    '1',
    '1', '1', '1', '1',
    '1', '1', '1',
    '1', '1',
]

URLPATH_REPLACABLES = [
    'API_VERSION',
    'USERNAME', 'USER_ID', 'PROVIDER_ID',
    'BANK_ID', 'BRANCH_ID', 'ATM_ID', 'PRODUCT_CODE',
    'OTHER_ACCOUNT_ID',
    'ACCOUNT_ID', 'VIEW_ID', 'TRANSACTION_ID', 'COUNTERPARTY_ID',
    'CUSTOMER_ID', 'MEETING_ID', 'CONSUMER_ID',
    'FROM_CURRENCY_CODE', 'TO_CURRENCY_CODE',
]

class IndexView(LoginRequiredMixin, TemplateView):
    """Index view for runtests"""
    template_name = "runtests/index.html"

    def get_testconfigs(self, testconfig_pk):
        testconfigs = {
            'available': [],
            'selected': None,
        }
        testconfigs['available'] = TestConfiguration.objects.filter(
            owner=self.request.user).order_by('name')
        if testconfig_pk:
            try:
                testconfigs['selected'] = TestConfiguration.objects.get(
                    owner=self.request.user,
                    pk=testconfig_pk,
                )
            except TestConfiguration.DoesNotExist as err:
                raise PermissionDenied
        return testconfigs

    def get_post_or_update(self, method, testconfigs, testconfig_pk, path, data, swagger ):

        params = ''
        order = 100
        # Get saved profile operations
        try:
            objs = ProfileOperation.objects.filter(
                profile_id=testconfig_pk,
                operation_id=data[method]['operationId']
            )
        except ProfileOperation.DoesNotExist:
            objs = None

        request_body = {}
        urlpath = self.get_urlpath(testconfigs["selected"], path)

        if objs is not None and len(objs)>0:
            objs_list = []
            for obj in objs:
                if obj.is_deleted==0:
                    params = obj.json_body
                    order = obj.order
                    urlpath = obj.urlpath
                    replica_id = obj.replica_id
                    remark = obj.remark if obj.remark is not None else data[method]['summary']
                    objs_list.append({
                        'urlpath': urlpath,
                        'method': method,
                        'order': order,
                        'params': params,
                        'summary': remark,
                        'operationId': data[method]['operationId'],
                        'replica_id':replica_id,
                        'responseCode': 200,
                    })
            return objs_list

        elif method == 'post' or method == 'put':
            # generate json body from swagger
            definition = data[method]['parameters'][0] if len(data[method]['parameters']) > 0 else None
            definition = definition['schema']['$ref'][14:]
            params = swagger['definitions'][definition]
            if len(params["required"]) > 0:
                for field in params["required"]:
                    # Match Profile variables
                    field_names = [ f.name for f in TestConfiguration._meta.fields]
                    if field in field_names:
                        request_body[field] = getattr(testconfigs["selected"], field)
                    else:
                        try:
                            request_body[field] = params["properties"][field].get("example", "")
                        except:
                            request_body[field] = None
            params = json.dumps(request_body, indent=4)

        return [{
            'urlpath': urlpath,
            'method': method,
            'order': order,
            'params': params,
            'summary': data[method]['summary'],
            'operationId': data[method]['operationId'],
            'replica_id':1,
            'responseCode': 200,
        }]

    def api_replace(self, string, match, value):
        """Helper to replace format strings from the API"""
        # API sometimes uses '{match}' or 'match' to denote variables
        return string. \
            replace('{{{}}}'.format(match), value). \
            replace(match, value)

    def get_urlpath(self, testconfig, path):
        """
        Gets a URL path
        where placeholders in given path are replaced by values from testconfig
        """
        urlpath = path
        for (index, match) in enumerate(URLPATH_REPLACABLES):
            value = getattr(testconfig, match.lower())
            if value:
                urlpath = self.api_replace(urlpath, match, value)
            else:
                urlpath = self.api_replace(urlpath, match, URLPATH_DEFAULT[index])

        return urlpath

    def get_context_data(self, **kwargs):
        context = super(IndexView, self).get_context_data(**kwargs)

        testconfig_pk = kwargs.get('testconfig_pk', 0)
        testconfigs = self.get_testconfigs(testconfig_pk)

        mappings = {
            'API_HOST':settings.API_HOST,
            'API_VERSION':settings.API_VERSION,
            'REDIRECT_URL':settings.REDIRECT_URL,
            'ADMIN_USERNAME':settings.ADMIN_USERNAME,
            'ADMIN_PASSWORD':settings.ADMIN_PASSWORD,
            'FILE_ROOT':settings.FILE_ROOT,
            'OUTPUT_PATH':settings.OUTPUT_PATH,
            'INPUT_PATH':settings.INPUT_PATH,
            'DATASET_PATH':settings.DATASET_PATH,
            'OPTIONS_PATH':settings.OPTIONS_PATH
        }

        context.update({
            'testconfigs': testconfigs,
            'testconfig_pk': testconfig_pk,
            'mappings': mappings
        })

        return context


class RunView(LoginRequiredMixin, TemplateView):
    """Run an actual test against the API"""
    template_name = "runtests/index.html"

    def render_to_response(self, context, **response_kwargs):
        return JsonResponse(self.get_data(context), **response_kwargs)

    def get_data(self, context):
        # This should ensure everything in context is JSON-serialisable
        if 'view' in context:
            del context['view']
        return context

    def api_replace(self, string, match, value):
        """Helper to replace format strings from the API"""
        # API sometimes uses '{match}' or 'match' to denote variables
        return string.\
            replace('{{{}}}'.format(match), value).\
            replace(match, value)

    def get_urlpath(self, testconfig, path):
        """
        Gets a URL path
        where placeholders in given path are replaced by values from testconfig
        """
        urlpath = path
        for index, match in enumerate(URLPATH_REPLACABLES):
            value = getattr(testconfig, match.lower())
            if value:
                urlpath = self.api_replace(urlpath, match, value)
            else:
                urlpath = self.api_replace(urlpath, match, URLPATH_DEFAULT[index])
        return urlpath

    def get_config(self, testmethod, testpath, testconfig_pk, operation_id):
        """Gets test config from swagger and database"""
        urlpath = urllib.parse.unquote(testpath)

        status_code = 200

        if testmethod == 'post':
            status_code = 201
        elif testmethod == 'put':
            status_code = 200
        elif testmethod == 'delete':
            status_code = 204

        try:
            objs = ProfileOperation.objects.filter(
                profile_id=int(testconfig_pk),
                operation_id=operation_id,
                is_deleted=0
            )
        except ProfileOperation.DoesNotExist:
            objs = None

        config = {
            'found': True,
            'method': testmethod,
            'status_code': status_code,
            'summary': 'Unknown',
            'urlpath': urlpath if objs is None or len(objs)==0 else objs[0].urlpath,
            'operation_id': operation_id,
            'profile_id': testconfig_pk,
            'payload': self.request.POST.get('json_body')
        }
        try:
            testconfig = TestConfiguration.objects.get(
                owner=self.request.user, pk=testconfig_pk)
        except TestConfiguration.DoesNotExist as err:
            raise PermissionDenied
        try:
            swagger = self.api.get_swagger(testconfig.api_version)
        except APIError as err:
            messages.error(self.request, err)
        else:
            for path, data in swagger['paths'].items():
                if testmethod in data and data[testmethod]['operationId'] == operation_id :
                    config.update({
                        'found': True,
                        'operation_id': data[testmethod]['operationId'],
                        'summary': data[testmethod]['summary'],
                        'urlpath': self.get_urlpath(testconfig, path),
                    })
        return config

    def post(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        return self.render_to_response(context)

    def run_test(self, config):
        """Runs a test with given config"""
        url = '{}{}'.format(settings.API_HOST, config['urlpath'])
        # Let APIError bubble up
        if config['method'] == 'get' or config['method'] == 'delete':
            response = self.api.call(config['method'], url)
        else:
            try:
                response = self.api.call(config['method'], url, json.loads(config['payload']))
            except:
                response = self.api.call(config['method'], url)
        try:
            text = response.json()
        except json.decoder.JSONDecodeError as err:
            text = response.text
        text = json.dumps(
            text, sort_keys=True, indent=2, separators=(',', ': '))
        result = {
            'text': text,
            'execution_time': response.execution_time,
            'status_code': response.status_code,
        }
        return result

    def get_context_data(self, **kwargs):
        context = super(RunView, self).get_context_data(**kwargs)
        self.api = API(self.request.session.get('obp'))
        payload = self.request.POST.get('json_body')
        config = self.get_config(**kwargs)

        if context['testpath'] is not None:
            config.update({'urlpath': context['testpath']})
        if payload is not None:
            config.update({'payload':payload})
        context.update({
            'config': config,
            'text': None,
            'execution_time': -1,
            'messages': [],
            'success': False,
        })
        if not config['found']:
            msg = 'Unknown path {}!'.format(kwargs['testpath'])
            context['messages'].append(msg)
            return context

        try:
            result = self.run_test(config)
        except APIError as err:
            context['messages'].append(err)
            return context
        else:
            context.update(result)

        # Test if status code is as expected
        if result['status_code'] != config['status_code']:
            msg = 'Status code is {}, but expected {}!'.format(
                result['status_code'], config['status_code'])
            context['messages'].append(msg)
            return context

        context['success'] = True
        return context


class TestConfigurationCreateView(LoginRequiredMixin, CreateView):
    model = TestConfiguration
    form_class = TestConfigurationForm

    def form_valid(self, form):
        form.instance.owner = self.request.user
        return super(TestConfigurationCreateView, self).form_valid(form)

    def get_success_url(self):
        return reverse('runtests-index-testconfig', kwargs={
            'testconfig_pk': self.object.pk,
        })


class TestConfigurationUpdateView(LoginRequiredMixin, UpdateView):
    model = TestConfiguration
    form_class = TestConfigurationForm

    def get_object(self, **kwargs):
        object = super(TestConfigurationUpdateView, self).get_object(**kwargs)
        if self.request.user != object.owner:
            raise PermissionDenied
        return object


    def form_valid(self, form):
        form.instance.owner = self.request.user
        return super(TestConfigurationUpdateView, self).form_valid(form)

    def get_success_url(self):
        return reverse('runtests-index-testconfig', kwargs={
            'testconfig_pk': self.object.pk,
        })


class TestConfigurationDeleteView(LoginRequiredMixin, DeleteView):
    model = TestConfiguration
    form_class = TestConfigurationForm
    success_url = reverse_lazy('runtests-index')

    def get_object(self, **kwargs):
        object = super(TestConfigurationDeleteView, self).get_object(**kwargs)
        if self.request.user != object.owner:
            raise PermissionDenied
        return object


def saveJsonBody(request):

    operation_id = request.POST.get('operation_id')
    json_body = request.POST.get('json_body', '')
    profile_id = request.POST.get('profile_id')
    order = request.POST.get('order')
    urlpath = request.POST.get('urlpath')
    replica_id = request.POST.get('replica_id')
    remark = request.POST.get('remark')

    #if not re.match("^{.*}$", json_body):
    #    json_body = "{{{}}}".format(json_body)

    data = {
        'operation_id' : operation_id,
        'json_body': json_body,
        'profile_id': profile_id,
        'order': order,
        'urlpath': urlpath,
        'remark':remark,
        'is_deleted':0
    }

    profile_list = ProfileOperation.objects.update_or_create(
        operation_id=operation_id,
        profile_id=profile_id,
        replica_id=replica_id,
        defaults=data
    )

    return JsonResponse({'state': True})


def copyJsonBody(request):
    saveJsonBody(request)

    operation_id = request.POST.get('operation_id')
    json_body = request.POST.get('json_body', '')
    profile_id = request.POST.get('profile_id')
    order = request.POST.get('order')
    urlpath = request.POST.get('urlpath')
    remark = request.POST.get('remark')

    #if not re.match("^{.*}$", json_body):
    #    json_body = "{{{}}}".format(json_body)

    profile_list = ProfileOperation.objects.filter(
        operation_id=operation_id,
        profile_id=profile_id
    )

    replica_id = max([profile.replica_id for profile in profile_list])+1

    ProfileOperation.objects.create(profile_id = profile_id, operation_id = operation_id, json_body = json_body, order = order, urlpath = urlpath, remark=remark, replica_id = replica_id, is_deleted=0)

    return JsonResponse({'state': True})

def deleteJsonBody(request):
    saveJsonBody(request)

    operation_id = request.POST.get('operation_id')
    profile_id = request.POST.get('profile_id')
    replica_id = request.POST.get('replica_id')

    profile = ProfileOperation.objects.get(
        operation_id=operation_id,
        profile_id=profile_id,
        replica_id=replica_id
    )
    profile.is_deleted = 1
    profile.save()

    return JsonResponse({'state': True})


class UserCreateView(LoginRequiredMixin, CreateView):
    model = User
    form_class = UserForm

    def form_valid(self, form):
        form.instance.owner = self.request.user
        return super(UserCreateView, self).form_valid(form)

    def get_success_url(self):
        return reverse('user-index', kwargs={
            'testconfig_pk': self.object.pk,
        })

class BankCreateView(LoginRequiredMixin, CreateView):
    model = Bank
    form_class = BankForm

    def form_valid(self, form):
        form.instance.owner = self.request.user
        return super(BankCreateView, self).form_valid(form)

    def get_success_url(self):
        return reverse('bank-index', kwargs={
            'testconfig_pk': self.object.pk,
        })

class BranchCreateView(LoginRequiredMixin, CreateView):
    model = Branch
    form_class = BranchForm

    def form_valid(self, form):
        form.instance.owner = self.request.user
        return super(BankCreateView, self).form_valid(form)

    def get_success_url(self):
        return reverse('bank-index', kwargs={
            'testconfig_pk': self.object.pk,
        })

class DataImportView(LoginRequiredMixin, TemplateView):
    template_name = "runtests/data_import.html"

    def get_context_data(self, **kwargs):
        context = super(DataImportView, self).get_context_data(**kwargs)
        mappings = {
            'API_HOST':settings.API_HOST,
            'ADMIN_USERNAME':settings.ADMIN_USERNAME,
            'ADMIN_PASSWORD':settings.ADMIN_PASSWORD,
            'OAUTH_CONSUMER_KEY':settings.OAUTH_CONSUMER_KEY,
        }

        context.update({
            'mappings': mappings
        })
        return context

class GenerateJsonView(LoginRequiredMixin, TemplateView):
    template_name = "runtests/generate_json.html"

    def get_context_data(self, **kwargs):
        context = super(GenerateJsonView, self).get_context_data(**kwargs)
        mappings = {
            'COUNTRY':settings.COUNTRY,
            'PRODUCT_NUMBER':settings.PRODUCT_NUMBER,
            'ATM_NUMBER':settings.ATM_NUMBER,
            'BRANCH_NUMBER':settings.BRANCH_NUMBER,
            'BANK_NUMBER':settings.BANK_NUMBER,
            'INPUT_PATH':settings.INPUT_PATH,
            'MONTHS':settings.MONTHS,
            'DATASET_PATH':settings.DATASET_PATH,
            'OPTIONS_PATH':settings.OPTIONS_PATH,
            'OUTPUT_PATH':settings.OUTPUT_PATH
        }

        context.update({
            'mappings': mappings
        })
        return context

def GenerateFile(request):
    from objects.User import User
    from objects.Bank import Bank
    from objects.Branch import ATM, Branch
    from objects.Product import Product
    country = request.POST.get('country')
    bank_number = int(request.POST.get('bank_num'))
    branch_number = int(request.POST.get('branch_num'))
    product_number = int(request.POST.get('product_num'))
    atm_number = int(request.POST.get('atm_num'))
    months = int(request.POST.get('months'))
    input_path = request.POST.get('input_path')

    output_path = request.POST.get('output_path')

    user_list = User.generator_for_file("{}options.xlsx".format(input_path))

    branch_list = []
    atm_list = []
    product_list = []
    bank_list = Bank.generate_from_file(bank_number)
    for bank in bank_list:
        branch_list.extend(Branch.generate_from_file(bank, branch_number))
        atm_list.extend(ATM.generate_from_file(bank, atm_number))
        product_list.extend(Product.generate_from_file(bank, product_number))

    account_list = []
    for user in user_list:
        branch_tmp = random.choice(branch_list)
        account1 = user.create_account(branch_tmp, "CURRENT", user.country, user.current)
        account1.set_behavior(income=user.current, spending_frequency =
        { "food":          user.food,
          "utility":       user.utility,
          "clothing":      user.clothing,
          "auto":          user.auto,
          "health":        user.health,
          "entertainment": user.entertainment,
          "gift":         user.gift,
          "education":     user.education,
          "fee":           user.fee },
                              housing_type={
                                  "RENT": -user.rent
                              } if user.rent > 0 else None
                              )
        account_list.append(account1)

        branch_tmp = random.choice(branch_list)
        account2 = user.create_account(branch_tmp, "SAVING", user.country, user.savings)
        account2.set_behavior(income=user.savings)
        account_list.append(account2)

    from datetime import date
    from datetime import timedelta
    date_start = date.today() - timedelta(days=months * 30)
    transaction_list = []
    for i in range(months):
        date_start = date_start + timedelta(days=30)
        for account in account_list:
            transaction_list.extend(account.generateTransaction(date_start))

    try:
        import os
        os.makedirs(output_path)
    except:
        pass
    with open('{}sandbox_pretty.json'.format(output_path), 'w') as outfile:
        json.dump({
            "users": user_list,
            "banks": bank_list,
            "branches": branch_list,
            "accounts": account_list,
            "atms":atm_list,
            # "counterparties":counterparty_list,
            "products":product_list,
            "transactions": transaction_list
        }, outfile, default=lambda x: x.dict(), indent=4)

    customer_list = []
    for user in user_list:
        for bank in bank_list:
            customer_list.append(user.create_customer(bank))

    with open('{}customers_pretty.json'.format(output_path), 'w') as outfile:
        json.dump(customer_list, outfile, default=lambda x: x.dict(), indent=4)

    wb = load_workbook("{}counterparties.xlsx".format(input_path))

    sheet_selected = [sheetname for sheetname in wb.sheetnames if country in sheetname ]

    dataframe = pd.read_excel("{}counterparties.xlsx".format(input_path), sheet_name = sheet_selected, header=None, index_col=None, skiprows=3)

    df = pd.DataFrame()
    for i, j in dataframe.items():
        df = pd.DataFrame(j)

    df.columns= ['type', 'name', 'reference', 'value', 'frequency', 'logo','homepage']

    df.type = df["type"].apply(lambda x: re.split('_|/',x)[0].lower())
    #df.frequency = df["frequency"].apply(lambda x: str("0") if x is np.nan else frequecy_mapping[x.lower()])

    df = pd.DataFrame({
        "type":df["type"].apply(lambda x: re.split('_|/',x)[0].lower()),
        "name" : df["name"],
        #"value" : df["value"],
        #"frequency" : df["frequency"].apply(lambda x: str("0") if x is np.nan else frequecy_mapping[x.lower()]),
        "logo" : df["logo"],
        "homepage" : df["homepage"]
    })

    df_list = []
    for i, rows in df.iterrows():
        df_list.append({
            "name":rows['name'],
            "category":rows['type'],
            "superCategory":rows['type'],
            "logoUrl":rows['logo'] if rows['logo'] is not np.nan else "",
            "homePageUrl":rows['homepage'] if rows['homepage'] is not np.nan else "",
            "region":country
        })

    with open('{}counterparty_pretty.json'.format(settings.OUTPUT_PATH), 'w') as outfile:
        json.dump(df_list, outfile, default=lambda x: x.dict(), indent=4)
    return JsonResponse({'state': True})

def ImportJson(request):
    api_host = request.POST.get('api_host')
    consumer_key = request.POST.get('consumer_key')
    adminUserUsername = request.POST.get('username')
    adminPassword = request.POST.get('password')

    from object.Admin import Admin
    admin_user = Admin(adminUserUsername, adminPassword, consumer_key)
    session = admin_user.direct_login(api_host)

    file_json = Admin.load(settings.FILE_ROOT+"sandbox_pretty.json")
    url = api_host+"/obp/v3.0.0/sandbox/data-import"
    result2 = session.request('POST', url, json=file_json, verify=settings.VERIFY)
    if result2.status_code==201:
        print("Update Successfully")
        return JsonResponse({'state': True})
    else:
        print("Update Failed")
        return JsonResponse({'state': False})

def ImportCounterparty(request):
    api_host = request.POST.get('api_host')
    consumer_key = request.POST.get('consumer_key')

    from objects.Admin import Admin
    from objects.PostCounterparty import PostCounterparty
    json_object_counterparty= PostCounterparty.load(settings.FILE_ROOT+"counterparty_pretty.json")

    counterparty_list = [val for sublist in json_object_counterparty for val in sublist]

    json_object_user=Admin.load(settings.FILE_ROOT+"sandbox_pretty.json")

    for user_dict in json_object_user['users']:
        user = Admin(user_dict['user_name'], user_dict['password'], consumer_key)
        print("login as user: ")
        session = user.direct_login(api_host)
        print("get users private accounts")
        private_account = user.get_user_private_account(api_host)
        account_list = json.loads(private_account)['accounts']
        print("ok!!!")

        print("get other accounts for the accounts")
        for account in account_list:
            bank_id = account['bank_id']
            region = bank_id.split('.')[2]
            account_id = account['id']
            view = account['views_available'][0]
            result = user.get_user_other_account(bank_id, account_id, view['id'])
            print(type(result))
            other_accounts_list = json.loads(result)['other_accounts']

            print("bank_id: {}".format(bank_id))

            print("region is {}".format(region))
            print("get matching json counterparty data for each transaction's other_account")
            for other_accounts in other_accounts_list:
                counterparty_name = other_accounts['holder']['name']
                print("Filtering counterparties by region {} and counterparty name {}".format(region, counterparty_name))
                regionCounterparties = [counterparty for counterparty in counterparty_list if counterparty['region']==region]
                records = [counterparty for counterparty in counterparty_list if counterparty['name'].lower()==counterparty_name.lower()]
                print("Found {} records".format(len(records)))
                for cp in records:
                    print("couterparty is Region {} Name {} Home Page {}".format(cp['region'],cp['name'],cp['homePageUrl']))
                    logoUrl = cp['homePageUrl'] if ("http://www.brandprofiles.com" in cp['logoUrl']) else cp['logoUrl']
                    if logoUrl.startswith("http") and other_accounts['metadata']['image_URL'] is None:
                        json_tmp = {"image_URL": logoUrl}
                        url = api_host + "/obp/v3.1.0/banks/" + bank_id + "/accounts/" + account_id + "/" + view['id'] + "/other_accounts/"+other_accounts['id']+"/metadata/image_url"
                        result = session.request('POST', url, json = json_tmp, verify=settings.VERIFY)
                        if result.status_code == 201:
                            print("Saved " + logoUrl + " as imageURL for counterparty "+ other_accounts['id'])
                        else:
                            print("Save failed. {}".format(result.error if result is not None and result.error is not None else ""))
                    else:
                        print("did NOT save " + logoUrl + " as imageURL for counterparty "+ other_accounts['id'])

                    if (cp['homePageUrl'].startswith("http") and not cp['homePageUrl'].endswith("jpg") and not cp['homePageUrl'].endswith("png") and other_accounts['metadata']['URL'] is None):
                        json_tmp = {"URL": cp['homePageUrl']}
                        url = api_host + "/obp/v3.1.0/banks/" + bank_id + "/accounts/" + account_id + "/" + \
                              view['id'] + "/other_accounts/" + other_accounts['id'] + "/metadata/url"
                        result = session.request('POST', url, json=json_tmp, verify=settings.VERIFY)
                        if result.status_code == 201:
                            print("Saved " + cp['homePageUrl'] + " as URL for counterparty "+ other_accounts['id'])
                        else:
                            print("Save failed. {}".format(result.error if result is not None and result.error is not None else ""))
                    else:
                        print("did NOT save " + cp['homePageUrl'] + " as URL for counterparty "+ other_accounts['id'])

                    if (cp['category'] is not None and other_accounts['metadata']['more_info'] is None):
                        categoryBits = cp['category'].split("_")
                        moreInfo = categoryBits[0]

                        json_tmp = {"more_info": moreInfo}
                        url = api_host + "/obp/v3.1.0/banks/" + bank_id + "/accounts/" + account_id + "/" + view['id'] + "/other_accounts/" + other_accounts['id'] + "/metadata/more_info"
                        result = session.request('POST', url, json=json_tmp, verify=settings.VERIFY)
                        if result.status_code==201:
                            print("Saved " + moreInfo + " as more_info for counterparty "+ other_accounts['id'])
                        else:
                            print("Save failed. {}".format(result.error if result is not None and result.error is not None else ""))
                    else:
                        if other_accounts['metadata']['more_info'] is not None:
                            print("more info is not empty:{}")
                        else:
                            print("did NOT save more_info for counterparty "+ other_accounts['id'])

        user.oauth_logout()
    return JsonResponse({'state': True})

def ImportCustomer(request):
    api_host = request.POST.get('api_host')
    consumer_key = request.POST.get('consumer_key')
    adminUserUsername = request.POST.get('username')
    adminPassword = request.POST.get('password')

    from object.PostCustomer import PostCustomer
    json_customers = PostCustomer.load(settings.FILE_ROOT+"/customers_pretty.json")

    print("Got {} records".format(len(json_customers)))

    customer_list = [PostCustomer(customer['customer_number'],
                                  customer['legal_name'],
                                  customer['mobile_phone_number'],
                                  customer['email'],
                                  customer['face_image'],
                                  customer['date_of_birth'],
                                  customer['relationship_status'],
                                  customer['dependants'],
                                  customer['dob_of_dependants'],
                                  customer['highest_education_attained'],
                                  customer['employment_status'],
                                  customer['kyc_status'],
                                  customer['last_ok_date'],
                                  customer['bank_id'],
                                  customer['credit_rating'],
                                  customer['credit_limit']
                                  ) for customer in json_customers]

    from objects.Admin import Admin
    from objects.Bank_Import import Bank_Import
    json_user = Admin.load(settings.FILE_ROOT+"sandbox_pretty.json")
    print("Got {} users".format(len(json_user['users'])))

    print("login as user: ")
    admin_user = Admin(adminUserUsername, adminPassword, consumer_key)
    session = admin_user.direct_login(api_host)
    print("login successfully!!!")

    print("Got {} banks".format(len(json_user['banks'])))
    bank_list=[]
    for bank in json_user['banks']:
        bank_list.append(Bank_Import(bank['id'],
                                     bank['short_name'],
                                     bank['full_name'],
                                     bank['logo'],
                                     bank['website']))

    for user_dict in json_user['users']:
        user = Admin(user_dict['user_name'], user_dict['password'], user_dict['email'])
        customer_filtered = [customer for customer in customer_list if customer.email == user.email]
        result = session.get(
            api_host + "/obp/v3.1.0/users/username/" + user.user_name)
        if result.status_code==200:
            current_user = json.loads(result.content)

            for customer in customer_filtered:
                print("email is {} customer number is {} name is {} and has {} dependants born on {} "
                           .format(customer.email, customer.customer_number, customer.legal_name, customer.dependants, customer.dob_of_dependants))

                for bank in bank_list:
                    print("Posting a customer for bank {}".format(bank.short_name))
                    url = api_host+"/obp/v2.1.0/banks/{}/customers".format(bank.id)
                    result2 = session.request('POST', url, json=customer.to_json(current_user['user_id']), verify=settings.VERIFY)
                    if result2.status_code==201:
                        print("saved {} as customer {}".format(customer.customer_number, result2.content))
                    else:
                        print("did NOT save customer {}".format(result2.content if result2 is not None and result2.content is not None else ""))
        else:
            print(result.content if result is not None and result.content is not None else "")
    return JsonResponse({'state': True})

class CounterpartyUpdateView(LoginRequiredMixin, TemplateView):
    template_name = "runtests/counterparty_update.html"

    def get_context_data(self, **kwargs):
        context = super(CounterpartyUpdateView, self).get_context_data(**kwargs)
        input_file=".\input_file\counterparties.xlsx"
        city='Hong Kong'
        wb = load_workbook(input_file)

        sheet_selected = [sheetname for sheetname in wb.sheetnames if city in sheetname ]

        import pandas as pd

        dataframe = pd.read_excel(input_file, sheet_name = sheet_selected, header=None, index_col=None, skiprows=3)

        df = pd.DataFrame()
        for i, j in dataframe.items():
            df = pd.DataFrame(j)

        df.columns= ['type', 'name', 'reference', 'value', 'frequency', 'logo','homepage']

        df.type = df["type"].apply(lambda x: re.split('_|/',x)[0].lower())
        #df.frequency = df["frequency"].apply(lambda x: str("0") if x is np.nan else frequecy_mapping[x.lower()])

        df = pd.DataFrame({
            "type":df["type"].apply(lambda x: re.split('_|/',x)[0].lower()),
            "name" : df["name"],
            #"value" : df["value"],
            #"frequency" : df["frequency"].apply(lambda x: str("0") if x is np.nan else frequecy_mapping[x.lower()]),
            "logo" : df["logo"],
            "homepage" : df["homepage"]
        })

        df_list = []
        for i, rows in df.iterrows():
            df_list.append({
                "name":rows['name'],
                "category":rows['type'],
                "superCategory":rows['type'],
                "logoUrl":rows['logo'] if rows['logo'] is not np.nan else "",
                "homePageUrl":rows['homepage'] if rows['homepage'] is not np.nan else "",
                "region":city
            })

        context.update({
            'calls': df_list
        })
        return context